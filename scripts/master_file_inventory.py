#!/usr/bin/env python3
"""
MASTER FILE INVENTORY
=====================
Analyzes EVERY file in EVERY zip across ALL years.
Documents exact structure: header rows, column names, data formats.
Creates a complete map before any parsing.
"""

import boto3
import pandas as pd
import zipfile
import re
import json
from io import BytesIO
from collections import defaultdict

S3_BUCKET = "ma-data123"
s3 = boto3.client('s3')

def list_all_raw_files():
    """List all raw files in S3."""
    files = []
    paginator = s3.get_paginator('list_objects_v2')
    for prefix in ['raw/stars/', 'raw/crosswalks/', 'raw/snp/', 'raw/enrollment/']:
        try:
            for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=prefix):
                for obj in page.get('Contents', []):
                    files.append({
                        'key': obj['Key'],
                        'size': obj['Size'],
                        'category': prefix.split('/')[1]
                    })
        except:
            pass
    return files

def extract_all_from_zip(zip_bytes: bytes) -> dict:
    """Extract all files from zip, handling nested zips."""
    files = {}
    try:
        zf = zipfile.ZipFile(BytesIO(zip_bytes))
        for name in zf.namelist():
            if name.endswith('.zip'):
                # Nested zip
                try:
                    inner_bytes = zf.read(name)
                    inner_zf = zipfile.ZipFile(BytesIO(inner_bytes))
                    for inner_name in inner_zf.namelist():
                        if inner_name.endswith(('.csv', '.xlsx', '.xls')):
                            files[f"{name}/{inner_name}"] = inner_zf.read(inner_name)
                except:
                    pass
            elif name.endswith(('.csv', '.xlsx', '.xls')):
                files[name] = zf.read(name)
    except Exception as e:
        pass
    return files

def analyze_csv_structure(content: bytes, filename: str) -> dict:
    """Analyze the structure of a CSV file - first 15 rows."""
    result = {
        'filename': filename,
        'file_type': 'csv',
        'encoding': None,
        'rows_analyzed': [],
        'likely_header_row': None,
        'likely_data_start': None,
        'column_count': 0,
        'has_contract_id': False,
        'has_measure_cols': False,
        'measure_col_format': None,
        'notes': []
    }

    for encoding in ['utf-8', 'latin-1', 'cp1252']:
        try:
            text = content.decode(encoding, errors='replace')
            result['encoding'] = encoding
            break
        except:
            continue

    if not result['encoding']:
        result['notes'].append('Could not decode file')
        return result

    lines = text.split('\n')[:15]

    for i, line in enumerate(lines):
        # Clean and split
        cols = line.strip().split(',')
        col_count = len([c for c in cols if c.strip()])

        # Analyze this row
        row_info = {
            'row_num': i,
            'col_count': col_count,
            'first_5_cols': [c.strip()[:50] for c in cols[:5]],
            'row_type': 'unknown'
        }

        line_lower = line.lower()

        # Detect row type
        if col_count <= 2 and ('star' in line_lower or 'rating' in line_lower or 'cms' in line_lower):
            row_info['row_type'] = 'title'
        elif 'contract' in line_lower and ('id' in line_lower or 'number' in line_lower):
            row_info['row_type'] = 'header'
            result['likely_header_row'] = i
            result['has_contract_id'] = True
        elif re.search(r'[CD]\d{2}[:\s]', line):
            row_info['row_type'] = 'measure_ids'
            result['has_measure_cols'] = True
            # Extract measure format
            measures = re.findall(r'([CD]\d{2})[:\s]([^,]+)', line)
            if measures:
                result['measure_col_format'] = measures[:3]  # First 3 examples
        elif re.search(r'^[HER]\d{4}', cols[0].strip()) if cols else False:
            row_info['row_type'] = 'data'
            if result['likely_data_start'] is None:
                result['likely_data_start'] = i
        elif 'domain' in line_lower or 'part c' in line_lower or 'part d' in line_lower:
            row_info['row_type'] = 'domain_header'
        elif re.search(r'\d{1,2}/\d{1,2}/\d{4}', line):
            row_info['row_type'] = 'date_row'
        elif '1 star' in line_lower or '2 star' in line_lower or 'star threshold' in line_lower:
            row_info['row_type'] = 'star_threshold'

        result['rows_analyzed'].append(row_info)
        result['column_count'] = max(result['column_count'], col_count)

    return result

def analyze_excel_structure(content: bytes, filename: str) -> dict:
    """Analyze Excel file structure."""
    result = {
        'filename': filename,
        'file_type': 'excel',
        'sheets': [],
        'notes': []
    }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
                row_data = [str(c)[:50] if c else '' for c in row[:5]]
                rows.append({'row_num': i, 'first_5_cols': row_data})
            result['sheets'].append({
                'name': sheet_name,
                'rows': rows
            })
    except Exception as e:
        result['notes'].append(f'Error reading Excel: {str(e)[:100]}')

    return result

def categorize_file(filename: str) -> str:
    """Categorize file by its name pattern."""
    fname_lower = filename.lower()

    if 'cut' in fname_lower and 'point' in fname_lower:
        return 'CUTPOINTS'
    elif 'measure' in fname_lower and 'star' in fname_lower:
        return 'MEASURE_STARS'
    elif 'measure' in fname_lower and 'data' in fname_lower:
        return 'MEASURE_DATA'
    elif 'summary' in fname_lower and 'rating' in fname_lower:
        return 'SUMMARY_RATING'
    elif 'domain' in fname_lower:
        return 'DOMAIN'
    elif 'cai' in fname_lower:
        return 'CAI'
    elif 'disenroll' in fname_lower:
        return 'DISENROLLMENT'
    elif 'display' in fname_lower:
        return 'DISPLAY_MEASURES'
    elif 'crosswalk' in fname_lower or 'xwalk' in fname_lower:
        return 'CROSSWALK'
    elif 'snp' in fname_lower:
        return 'SNP'
    elif 'enrollment' in fname_lower:
        return 'ENROLLMENT'
    else:
        return 'OTHER'

def extract_year(key: str, filename: str) -> int:
    """Extract year from path or filename."""
    # Try path first
    match = re.search(r'/(\d{4})_', key)
    if match:
        return int(match.group(1))

    # Try filename
    match = re.search(r'(\d{4})', filename)
    if match:
        year = int(match.group(1))
        if 2000 <= year <= 2030:
            return year

    return None

def main():
    print("=" * 80)
    print("MASTER FILE INVENTORY - Analyzing ALL files across ALL years")
    print("=" * 80)

    # Get all raw files
    raw_files = list_all_raw_files()
    print(f"\nFound {len(raw_files)} raw files in S3")

    # Master inventory
    inventory = []

    # Process each zip
    for rf in raw_files:
        if not rf['key'].endswith('.zip'):
            continue

        print(f"\nProcessing: {rf['key']}")

        try:
            resp = s3.get_object(Bucket=S3_BUCKET, Key=rf['key'])
            zip_bytes = resp['Body'].read()
            files = extract_all_from_zip(zip_bytes)

            print(f"  Contains {len(files)} files")

            for fname, content in files.items():
                year = extract_year(rf['key'], fname)
                category = categorize_file(fname)

                if fname.endswith('.csv'):
                    structure = analyze_csv_structure(content, fname)
                elif fname.endswith(('.xlsx', '.xls')):
                    structure = analyze_excel_structure(content, fname)
                else:
                    structure = {'filename': fname, 'notes': ['Unknown format']}

                inventory.append({
                    's3_key': rf['key'],
                    'inner_file': fname,
                    'year': year,
                    'category': category,
                    'size_bytes': len(content),
                    'structure': structure
                })

        except Exception as e:
            print(f"  Error: {e}")

    # Now create summary tables
    print("\n" + "=" * 80)
    print("MASTER INVENTORY SUMMARY")
    print("=" * 80)

    # Group by category and year
    by_category = defaultdict(list)
    for item in inventory:
        by_category[item['category']].append(item)

    # Print detailed summary for each category
    for category in sorted(by_category.keys()):
        items = by_category[category]
        years = sorted(set(i['year'] for i in items if i['year']))

        print(f"\n{'='*60}")
        print(f"CATEGORY: {category}")
        print(f"{'='*60}")
        print(f"Total files: {len(items)}")
        print(f"Years covered: {min(years) if years else 'N/A'} - {max(years) if years else 'N/A'}")
        print(f"Year list: {years}")

        # Show structure variations
        structures = defaultdict(list)
        for item in items:
            s = item['structure']
            if 'rows_analyzed' in s:
                # Create structure signature
                sig = []
                for row in s['rows_analyzed'][:6]:
                    sig.append(row.get('row_type', 'unknown'))
                sig_str = ' -> '.join(sig)
                structures[sig_str].append(item['year'])

        print(f"\nStructure variations:")
        for sig, sig_years in structures.items():
            print(f"  {sig}")
            print(f"    Years: {sorted(set(sig_years))}")

        # Show sample file details
        print(f"\nSample files:")
        seen_years = set()
        for item in sorted(items, key=lambda x: x['year'] or 0):
            if item['year'] and item['year'] not in seen_years:
                seen_years.add(item['year'])
                s = item['structure']
                print(f"\n  Year {item['year']}: {item['inner_file'].split('/')[-1]}")
                if 'rows_analyzed' in s:
                    for row in s['rows_analyzed'][:6]:
                        print(f"    Row {row['row_num']}: [{row['row_type']}] {row['first_5_cols']}")
                if len(seen_years) >= 5:  # Show 5 sample years per category
                    break

    # Save full inventory to JSON
    output_path = '/tmp/master_file_inventory.json'
    with open(output_path, 'w') as f:
        json.dump(inventory, f, indent=2, default=str)
    print(f"\n\nFull inventory saved to: {output_path}")

    # Create CSV summary
    summary_rows = []
    for item in inventory:
        s = item['structure']
        row_types = [r.get('row_type', '') for r in s.get('rows_analyzed', [])[:6]]
        summary_rows.append({
            'category': item['category'],
            'year': item['year'],
            's3_key': item['s3_key'],
            'inner_file': item['inner_file'].split('/')[-1],
            'size_kb': round(item['size_bytes'] / 1024, 1),
            'encoding': s.get('encoding', ''),
            'col_count': s.get('column_count', 0),
            'header_row': s.get('likely_header_row', ''),
            'data_start': s.get('likely_data_start', ''),
            'has_contract': s.get('has_contract_id', False),
            'has_measures': s.get('has_measure_cols', False),
            'row_0_type': row_types[0] if len(row_types) > 0 else '',
            'row_1_type': row_types[1] if len(row_types) > 1 else '',
            'row_2_type': row_types[2] if len(row_types) > 2 else '',
            'row_3_type': row_types[3] if len(row_types) > 3 else '',
            'row_4_type': row_types[4] if len(row_types) > 4 else '',
        })

    summary_df = pd.DataFrame(summary_rows)
    summary_df = summary_df.sort_values(['category', 'year'])
    summary_csv = '/tmp/master_file_inventory.csv'
    summary_df.to_csv(summary_csv, index=False)
    print(f"CSV summary saved to: {summary_csv}")

    # Print the CSV for easy viewing
    print("\n" + "=" * 80)
    print("COMPLETE FILE INVENTORY TABLE")
    print("=" * 80)
    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    print(summary_df.to_string(index=False))

if __name__ == '__main__':
    main()
